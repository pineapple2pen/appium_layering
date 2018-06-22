# -*- coding: utf-8 -*-
# @Author : Pineapple Dong ^_^
# @Time   : 2018/6/21 19:37
# @File   : excel_tool.py
import openpyxl
from openpyxl.worksheet.worksheet import Worksheet as XlsxSheet
import os


class ExcelTool(object):
    work_book_xlsx = None
    load_file = False

    def __init__(self, operator_type, file_path):
        """
        生成excel文件,建议使用xlsx格式，07及以上office默认文件为xlsx格式
        :param operator_type: 操作类型(w:新建文件读写;r:导入文件读写)
        :param file_path 文件保存路径
        """
        if operator_type == "w" and file_path:
            self.file_name = os.path.basename(file_path)
            self.check_file_name(self.file_name)
            self.work_book_xlsx = openpyxl.Workbook()
            self.file_path = os.path.dirname(file_path)
        elif operator_type == "r" and file_path:
            self.load_file = True
            self.work_book_xlsx = openpyxl.load_workbook(file_path)
        else:
            raise Exception("Path must not empty or operator type error!")

    def add_xlsx_sheet(self, sheet_name: str, index: int = 0, no_insert: bool = True) -> XlsxSheet:
        """
        返回excel.xlsx的Sheet对象
        注意：插入值得方法为 sheet.cell(row=1,column=1).value = "value" 或者 sheet.cell(row=1,column=1,value="value")
        :param sheet_name: sheet名称
        :param index: sheet编号
        :param no_insert:新建的sheet是否插入至sheet编号位置
        :return: xlsx.Sheet对象
        """
        if index == 0 and no_insert:
            return self.work_book_xlsx.active
        return self.work_book_xlsx.create_sheet(sheet_name, index)

    def save_file(self, file_name=None):

        if file_name:
            self.check_file_name(file_name)
            self.work_book_xlsx.save(file_name)
        elif self.file_name:
            self.work_book_xlsx.save(os.path.join(self.file_path, self.file_name))
        else:
            raise NotFileNameError("not set file name!")

    @staticmethod
    def check_file_format(file_format):
        file_formats = "xlsx"
        if file_format != file_formats:
            raise TypeError

    def check_file_name(self, file_name: str):
        file_format = file_name.split(".")[-1]
        self.check_file_format(file_format)


class NotFileNameError(Exception):
    def __init__(self, describe):
        err = "Error because {}".format(describe)
        Exception.__init__(self, err)
        self.describe = describe


if __name__ == "__main__":
    test_case = ExcelTool("r", "../TestCase.xlsx")
    sheet1 = test_case.work_book_xlsx["用例"]
    print(sheet1.cell(1, 1).value)

    test_case.work_book_xlsx.close()
